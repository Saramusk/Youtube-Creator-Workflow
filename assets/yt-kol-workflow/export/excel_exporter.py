#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Local Excel exports aligned with the canonical Feishu table schemas."""

from __future__ import annotations

import logging
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from feishu.schema import (
    INFLUENCERS_FIELDS,
    INFLUENCER_VIDEOS_FIELDS,
    PRIMARY_FIELD_NAME,
    SEARCH_TASKS_FIELDS,
    SEARCH_VIDEOS_FIELDS,
)

logger = logging.getLogger("kol_workflow.export.excel_exporter")

# Header styling
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style="thin", color="2F5496"),
    right=Side(style="thin", color="D9E2F3"),
)


# ============================================================================
# Canonical column mappings: internal source field -> Feishu field name
#
# Output names and order are always derived from ``feishu.schema``.  The alias
# dictionaries below only describe how workflow-internal objects expose each
# value; they are not a second copy of the output schema.
# ============================================================================

_BLANK_PRIMARY_SOURCE = "__blank_feishu_primary_field__"


def _columns_from_schema(
    field_specs: Sequence[Tuple[str, int]],
    source_aliases: Mapping[str, str],
) -> Dict[str, str]:
    columns: Dict[str, str] = {}
    for field_name, _field_type in field_specs:
        source_name = source_aliases.get(field_name, field_name)
        if source_name in columns:
            raise ValueError(f"Excel 字段来源键重复: {source_name}")
        columns[source_name] = field_name
    return columns


SEARCH_TASK_COLUMNS = _columns_from_schema(
    SEARCH_TASKS_FIELDS,
    {
        PRIMARY_FIELD_NAME: _BLANK_PRIMARY_SOURCE,
        "唯一键": "task_key",
        "搜索关键词": "keyword",
        "排序策略": "sort_order",
        "地区": "region",
        "搜索时间": "search_time",
        "搜索结果数": "result_count",
        "筛选通过数": "qualified_count",
        "独立频道数": "unique_channels",
        "新增网红数": "new_channels",
        "配额消耗": "quota_used",
        "执行状态": "status",
        "备注": "note",
    },
)

SEARCH_VIDEO_COLUMNS = _columns_from_schema(
    SEARCH_VIDEOS_FIELDS,
    {
        PRIMARY_FIELD_NAME: _BLANK_PRIMARY_SOURCE,
        "唯一键": "search_video_unique_key",
        "视频记录日期": "video_record_date",
        "搜索关键词": "搜索关键词",
        "视频URL": "video_url",
        "Video ID": "video_id",
        "Channel ID": "channel_id",
        "Channel Name": "channel_title",
        "Video Title": "title",
        "Publish Time": "published_at",
        "Views": "view_count",
        "Likes": "like_count",
        "Comments": "comment_count",
        "互动率(%)": "engagement_rate",
        "Duration (sec)": "duration_seconds",
        "Duration (H:M:S)": "duration_hms",
        "Tags": "tags",
        "Has Subtitles": "has_caption",
        "是否通过筛选": "is_qualified",
        "筛选原因": "filter_reason",
    },
)

INFLUENCER_COLUMNS = _columns_from_schema(
    INFLUENCERS_FIELDS,
    {
        PRIMARY_FIELD_NAME: _BLANK_PRIMARY_SOURCE,
        "Channel ID": "channel_id",
        "网红记录日期": "influencer_record_date",
        "Channel Name": "channel_title",
        "KOL Name": "kol_name",
        "最新发布日期": "latest_published_at",
        "断更评估": "activity_status",
        "频道URL": "channel_url",
        "订阅数": "subscriber_count",
        "频道总播放量": "total_view_count",
        "视频总数": "total_video_count",
        "国家/地区": "country",
        "频道创建日期": "channel_created_at",
        "频道描述": "channel_description",
        "频道初步判断": "channel_initial_assessment",
        "联系邮箱": "contact_email",
        "邮箱状态": "email_status",
        "代表视频URL": "rep_video_url",
        "代表视频标题": "rep_video_title",
        "代表视频播放量": "rep_video_views",
        "代表视频互动率": "rep_video_engagement",
        "来源关键词": "source_keyword",
        "开发状态": "dev_status",
    },
)

INFLUENCER_VIDEO_COLUMNS = _columns_from_schema(
    INFLUENCER_VIDEOS_FIELDS,
    {
        PRIMARY_FIELD_NAME: _BLANK_PRIMARY_SOURCE,
        "唯一键": "influencer_video_unique_key",
        "Channel ID": "channel_id",
        "Channel Name": "channel_title",
        "视频URL": "video_url",
        "Video ID": "video_id",
        "Video Title": "title",
        "Publish Time": "published_at",
        "Views": "view_count",
        "Likes": "like_count",
        "Comments": "comment_count",
        "互动率(%)": "engagement_rate",
        "Duration (sec)": "duration_seconds",
        "Duration (H:M:S)": "duration_hms",
        "Tags": "tags",
        "字幕内容": "subtitle_content",
    },
)

TABLE_FIELD_TYPES: Dict[str, Dict[str, int]] = {
    "search_tasks": dict(SEARCH_TASKS_FIELDS),
    "search_videos": dict(SEARCH_VIDEOS_FIELDS),
    "influencers": dict(INFLUENCERS_FIELDS),
    "influencer_videos": dict(INFLUENCER_VIDEOS_FIELDS),
}

_FIELD_TYPES_BY_NAME: Dict[str, int] = {}
for _table_types in TABLE_FIELD_TYPES.values():
    for _field_name, _field_type in _table_types.items():
        previous = _FIELD_TYPES_BY_NAME.setdefault(_field_name, _field_type)
        if previous != _field_type:
            raise ValueError(f"同名飞书字段类型不一致: {_field_name}")


def _find_value(item: Mapping[str, Any], source: str, header: str) -> Any:
    """Read an internal value while also accepting schema-shaped records."""
    if header == PRIMARY_FIELD_NAME:
        return ""

    nested_fields = item.get("fields")
    containers = [item]
    if isinstance(nested_fields, Mapping):
        containers.append(nested_fields)

    blank_value: Any = ""
    for container in containers:
        if source in container:
            value = container.get(source)
            if not _is_blank(value):
                return value
            blank_value = value
        if header in container:
            value = container.get(header)
            if not _is_blank(value):
                return value
            blank_value = value

    # Composite keys are stored in Feishu but derived from the raw workflow
    # objects before local export.
    if source == "task_key":
        return item.get("keyword", "")
    if source == "search_video_unique_key":
        keyword = item.get("搜索关键词") or item.get("keyword") or ""
        video_id = item.get("video_id") or item.get("Video ID") or ""
        return f"{keyword}|{video_id}" if keyword or video_id else ""
    if source == "influencer_video_unique_key":
        return item.get("video_id") or item.get("Video ID") or ""
    return blank_value


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    try:
        result = pd.isna(value)
        return bool(result) if not isinstance(result, (list, tuple)) else False
    except (TypeError, ValueError):
        return False


def _plain_url(value: Any) -> str:
    if _is_blank(value):
        return ""
    if isinstance(value, Mapping):
        return str(value.get("link") or value.get("text") or "")
    return str(value)


def _excel_datetime(value: Any) -> Any:
    if _is_blank(value):
        return ""
    if isinstance(value, datetime):
        dt = value
        if dt.tzinfo is not None:
            dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
        return dt
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # Feishu date fields are Unix milliseconds.  Retain ordinary Excel
        # serial values if a caller supplied one rather than guessing.
        if abs(value) > 10_000_000_000:
            return datetime.fromtimestamp(value / 1000, tz=timezone.utc).replace(tzinfo=None)
        return value
    if isinstance(value, str):
        numeric_text = value.strip()
        try:
            numeric_value = float(numeric_text)
        except ValueError:
            pass
        else:
            if abs(numeric_value) > 10_000_000_000:
                return datetime.fromtimestamp(
                    numeric_value / 1000,
                    tz=timezone.utc,
                ).replace(tzinfo=None)
    try:
        parsed = pd.to_datetime(value, errors="coerce", utc=True)
    except (TypeError, ValueError):
        return value
    if pd.isna(parsed):
        return value
    return parsed.to_pydatetime().replace(tzinfo=None)


def _excel_number(value: Any) -> Any:
    if _is_blank(value):
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    text = str(value).strip().replace(",", "").replace("%", "")
    try:
        parsed = float(text)
    except ValueError:
        return value
    return int(parsed) if parsed.is_integer() else parsed


def _excel_bool(value: Any) -> Any:
    if _is_blank(value):
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    normalized = str(value).strip().lower()
    if normalized in {"true", "yes", "y", "1", "是", "有", "通过"}:
        return True
    if normalized in {"false", "no", "n", "0", "否", "无", "未通过"}:
        return False
    return value


def _coerce_excel_value(value: Any, field_type: Optional[int]) -> Any:
    if field_type == 15:
        return _plain_url(value)
    if field_type in {5, 1001}:
        return _excel_datetime(value)
    if field_type == 2:
        return _excel_number(value)
    if field_type == 7:
        return _excel_bool(value)
    if _is_blank(value):
        return ""
    if isinstance(value, Mapping):
        return str(value.get("text") or value.get("name") or value.get("link") or "")
    if isinstance(value, (list, tuple, set)):
        return ",".join(str(part) for part in value)
    return value


def export_to_excel(
    data: List[Dict],
    columns: Dict[str, str],
    filepath: str,
    sheet_name: str = "Sheet1",
    field_types: Optional[Mapping[str, int]] = None,
) -> str:
    """Export records to a formatted Excel file using the requested schema."""
    path = Path(filepath)
    path.parent.mkdir(parents=True, exist_ok=True)
    types = field_types or _FIELD_TYPES_BY_NAME

    rows = []
    for item in data:
        row = {}
        for source, header in columns.items():
            raw_value = _find_value(item, source, header)
            row[header] = _coerce_excel_value(raw_value, types.get(header))
        rows.append(row)

    df = pd.DataFrame(rows, columns=list(columns.values()))

    with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        _style_worksheet(ws, types)

    logger.info(f"Excel 导出: {path} ({len(data)} 条)")
    return str(path.resolve())


def _style_worksheet(ws, field_types: Optional[Mapping[str, int]] = None):
    """Apply header styling, types, hyperlinks, auto-width and freeze row 1."""
    types = field_types or _FIELD_TYPES_BY_NAME
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER

    ws.freeze_panes = "A2"

    for col_idx, col_cells in enumerate(ws.columns, 1):
        cells = list(col_cells)
        header = str(cells[0].value or "") if cells else ""
        field_type = types.get(header)
        max_length = 0
        for row_idx, cell in enumerate(cells):
            try:
                value = cell.value
                text = str(value or "")
                length = sum(2 if ord(char) > 127 else 1 for char in text)
                max_length = max(max_length, length)
                if row_idx and length > 60:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    wrapped_lines = min(4, max(2, (length + 59) // 60))
                    current_height = ws.row_dimensions[row_idx + 1].height or 15
                    ws.row_dimensions[row_idx + 1].height = max(
                        current_height,
                        wrapped_lines * 15,
                    )
                if row_idx and field_type in {5, 1001} and value not in {None, ""}:
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"
                if row_idx and field_type == 15 and isinstance(value, str) and value.startswith(("http://", "https://")):
                    cell.hyperlink = value
                    cell.style = "Hyperlink"
            except Exception:
                pass
        adjusted = min(max_length + 4, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(adjusted, 10)

    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


class ExcelExporter:
    """Manage the four Excel exports for one workflow run."""

    def __init__(self, output_dir: str, keyword: str = "", timestamp: str = ""):
        self.timestamp = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_kw = "".join(char if char.isalnum() or char in " _-" else "_" for char in keyword)
        safe_kw = safe_kw.strip().replace(" ", "_")[:50]
        self.base_dir = (
            Path(output_dir) / f"{self.timestamp}_{safe_kw}"
            if safe_kw
            else Path(output_dir) / self.timestamp
        )
        self.base_dir.mkdir(parents=True, exist_ok=True)
        self.files_created: List[str] = []

    def export_search_tasks(self, tasks: List[Dict]) -> str:
        """Export search-task audit records."""
        fp = str(self.base_dir / "search_tasks.xlsx")
        path = export_to_excel(
            tasks,
            SEARCH_TASK_COLUMNS,
            fp,
            "搜索任务",
            TABLE_FIELD_TYPES["search_tasks"],
        )
        self.files_created.append(path)
        return path

    def export_search_videos(self, videos: List[Dict], keyword: str = "") -> str:
        """Export search results, including their filter status."""
        # Preserve the historical side effect because batch collection happens
        # after the per-keyword export and relies on the source keyword.
        for video in videos:
            if keyword:
                video["搜索关键词"] = keyword
            video_id = video.get("video_id") or video.get("Video ID") or ""
            source_keyword = video.get("搜索关键词") or video.get("keyword") or ""
            if not video.get("search_video_unique_key"):
                video["search_video_unique_key"] = f"{source_keyword}|{video_id}"
        fp = str(self.base_dir / "search_videos.xlsx")
        path = export_to_excel(
            videos,
            SEARCH_VIDEO_COLUMNS,
            fp,
            "搜索结果视频",
            TABLE_FIELD_TYPES["search_videos"],
        )
        self.files_created.append(path)
        return path

    def export_influencers(self, influencers: List[Dict]) -> str:
        """Export influencer details."""
        fp = str(self.base_dir / "influencers.xlsx")
        path = export_to_excel(
            influencers,
            INFLUENCER_COLUMNS,
            fp,
            "网红详情",
            TABLE_FIELD_TYPES["influencers"],
        )
        self.files_created.append(path)
        return path

    def export_influencer_videos(self, videos: List[Dict]) -> str:
        """Export influencer recent videos."""
        fp = str(self.base_dir / "influencer_videos.xlsx")
        path = export_to_excel(
            videos,
            INFLUENCER_VIDEO_COLUMNS,
            fp,
            "网红视频",
            TABLE_FIELD_TYPES["influencer_videos"],
        )
        self.files_created.append(path)
        return path

    def ensure_all_files(self) -> List[str]:
        """Create header-only files for tables not exported during this run.

        Existing files are never overwritten, so a finalizer can safely call
        this after successful, empty, interrupted, or partially resumed runs.
        """
        specs = [
            (
                "search_tasks.xlsx",
                SEARCH_TASK_COLUMNS,
                "搜索任务",
                "search_tasks",
            ),
            (
                "search_videos.xlsx",
                SEARCH_VIDEO_COLUMNS,
                "搜索结果视频",
                "search_videos",
            ),
            (
                "influencers.xlsx",
                INFLUENCER_COLUMNS,
                "网红详情",
                "influencers",
            ),
            (
                "influencer_videos.xlsx",
                INFLUENCER_VIDEO_COLUMNS,
                "网红视频",
                "influencer_videos",
            ),
        ]
        paths: List[str] = []
        for filename, columns, sheet_name, table_key in specs:
            filepath = self.base_dir / filename
            if filepath.exists():
                path = str(filepath.resolve())
            else:
                path = export_to_excel(
                    [],
                    columns,
                    str(filepath),
                    sheet_name,
                    TABLE_FIELD_TYPES[table_key],
                )
            if path not in self.files_created:
                self.files_created.append(path)
            paths.append(path)
        return paths

    def get_output_dir(self) -> str:
        return str(self.base_dir)


class BatchExcelExporter:
    """Manage Excel exports for batch (multi-keyword) runs."""

    def __init__(self, output_dir: str):
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.base_dir = Path(output_dir) / f"{self.timestamp}_batch"
        self.base_dir.mkdir(parents=True, exist_ok=True)
        self.per_keyword_dir = self.base_dir / "per_keyword"
        self.per_keyword_dir.mkdir(exist_ok=True)

        self.all_search_tasks: List[Dict] = []
        self.all_search_videos: List[Dict] = []
        self.all_influencers: List[Dict] = []
        self.all_influencer_videos: List[Dict] = []
        self.files_created: List[str] = []

    def get_keyword_exporter(self, keyword: str) -> ExcelExporter:
        """Get an ExcelExporter for a keyword-specific subdirectory."""
        return ExcelExporter(
            str(self.per_keyword_dir),
            keyword=keyword,
            timestamp=self.timestamp,
        )

    def accumulate(
        self,
        search_tasks: Optional[List[Dict]] = None,
        search_videos: Optional[List[Dict]] = None,
        influencers: Optional[List[Dict]] = None,
        influencer_videos: Optional[List[Dict]] = None,
    ):
        """Accumulate four-table data across keywords for summary files."""
        if search_tasks:
            self.all_search_tasks.extend(search_tasks)
        if search_videos:
            self.all_search_videos.extend(search_videos)
        if influencers:
            self.all_influencers.extend(influencers)
        if influencer_videos:
            self.all_influencer_videos.extend(influencer_videos)

    def export_summary(self) -> List[str]:
        """Export four schema-complete aggregate files, including empty tables."""
        specs = [
            (
                "search_tasks_all.xlsx",
                self.all_search_tasks,
                SEARCH_TASK_COLUMNS,
                "全部搜索任务",
                "search_tasks",
            ),
            (
                "search_videos_all.xlsx",
                self.all_search_videos,
                SEARCH_VIDEO_COLUMNS,
                "全部视频",
                "search_videos",
            ),
            (
                "influencers_all.xlsx",
                self.all_influencers,
                INFLUENCER_COLUMNS,
                "全部网红",
                "influencers",
            ),
            (
                "influencer_videos_all.xlsx",
                self.all_influencer_videos,
                INFLUENCER_VIDEO_COLUMNS,
                "全部网红视频",
                "influencer_videos",
            ),
        ]
        for filename, rows, columns, sheet_name, table_key in specs:
            fp = str(self.base_dir / filename)
            path = export_to_excel(
                rows,
                columns,
                fp,
                sheet_name,
                TABLE_FIELD_TYPES[table_key],
            )
            self.files_created.append(path)
        return self.files_created

    def get_output_dir(self) -> str:
        return str(self.base_dir)
